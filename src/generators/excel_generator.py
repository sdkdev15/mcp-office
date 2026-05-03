"""Excel file generator using openpyxl with theme support and auto chart detection."""

from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.chart import Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.styles.themes import get_theme
from src.styles.style_applier import StyleApplier
from src.utils.data_transformer import DataTransformer
from src.utils.logger import get_logger
from src.utils.validators import validate_sheet_data, validate_chart_type, ValidationError

log = get_logger("excel_generator")
transformer = DataTransformer()


class ExcelGenerator:
    """Generates Excel workbooks with rich styling, charts, and formatting."""

    def __init__(self, theme_name: str = "corporate"):
        self.theme_name = theme_name
        self.theme = get_theme(theme_name)
        self.style_applier = StyleApplier(theme_name)

    def create_workbook(self, sheets: list[dict], metadata: Optional[dict] = None) -> bytes:
        """Create a complete Excel workbook from sheet data.

        Args:
            sheets: List of sheet data dictionaries with name, headers, rows.
            metadata: Optional document metadata.

        Returns:
            Workbook content as bytes.
        """
        validate_sheet_data(sheets)
        wb = Workbook()

        # Remove default sheet (openpyxl creates "Sheet" by default)
        if wb.sheetnames:
            del wb[wb.sheetnames[0]]

        # Apply metadata
        if metadata:
            self._apply_metadata(wb, metadata)

        for sheet_data in sheets:
            self._add_sheet(wb, sheet_data)

        return self._save_workbook(wb)

    def add_chart_to_sheet(
        self,
        wb: Workbook,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        title: str = "Chart",
        position: str = "D2",
    ) -> bytes:
        """Add a chart to an existing sheet.

        Args:
            wb: Workbook instance.
            sheet_name: Target sheet name.
            chart_type: Chart type (bar, line, pie, etc.).
            data_range: Cell range for data (e.g., 'A1:B10').
            title: Chart title.
            position: Cell position for chart.

        Returns:
            Updated workbook as bytes.
        """
        chart_type = validate_chart_type(chart_type)
        ws = wb[sheet_name]
        chart = self.style_applier.create_chart(chart_type, title)

        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=True)
        ws.add_chart(chart, position)

        return self._save_workbook(wb)

    def _add_sheet(self, wb: Workbook, sheet_data: dict) -> None:
        """Add a formatted sheet to the workbook.

        Args:
            wb: Workbook instance.
            sheet_data: Sheet data with name, headers, rows.
        """
        name = sheet_data.get("name", f"Sheet{len(wb.sheetnames) + 1}")
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        ws = wb.create_sheet(title=name)

        # Write headers
        if headers:
            header_row = []
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                header_row.append(cell)
            self.style_applier.apply_header_style(header_row, ws)

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            data_row = []
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
                data_row.append(ws.cell(row=row_idx, column=col_idx))
            is_alternate = (row_idx - 2) % 2 == 1
            self.style_applier.apply_data_style(data_row, ws, is_alternate)

        # Auto-fit columns
        self.style_applier.auto_fit_columns(ws)

    def _apply_metadata(self, wb: Workbook, metadata: dict) -> None:
        """Apply document metadata.

        Args:
            wb: Workbook instance.
            metadata: Metadata dictionary.
        """
        props = wb.properties
        if metadata.get("author"):
            props.creator = metadata["author"]
        if metadata.get("company"):
            props.company = metadata["company"]
        if metadata.get("subject"):
            props.subject = metadata["subject"]
        if metadata.get("title"):
            props.title = metadata["title"]
        if metadata.get("keywords"):
            props.keywords = metadata["keywords"]
        if metadata.get("category"):
            props.category = metadata["category"]
        if metadata.get("comments"):
            props.description = metadata["comments"]

    def _save_workbook(self, wb: Workbook) -> bytes:
        """Save workbook to bytes.

        Args:
            wb: Workbook instance.

        Returns:
            Workbook content as bytes.
        """
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def create_from_prompt(self, prompt: str, filename: str = "report.xlsx") -> bytes:
        """Create an Excel workbook from a natural language prompt.

        This is a simplified implementation — in production, this would use
        LLM parsing to extract structured data from the prompt.

        Args:
            prompt: Natural language description.
            filename: Output filename.

        Returns:
            Workbook content as bytes.
        """
        log.info(f"Creating Excel from prompt: {prompt[:100]}...")

        # Default template based on common patterns
        sheets = [
            {
                "name": "Summary",
                "headers": ["Item", "Value", "Status"],
                "rows": [
                    ["Generated from prompt", prompt[:50], "Complete"],
                ],
            }
        ]

        return self.create_workbook(sheets)

    def export_to_csv(self, sheets: list[dict]) -> dict[str, bytes]:
        """Export sheet data to CSV format.

        Args:
            sheets: List of sheet data.

        Returns:
            Dictionary mapping sheet names to CSV bytes.
        """
        result = {}
        for sheet_data in sheets:
            name = sheet_data.get("name", "Sheet")
            headers = sheet_data.get("headers", [])
            rows = sheet_data.get("rows", [])

            lines = []
            if headers:
                lines.append(",".join(f'"{h}"' for h in headers))
            for row in rows:
                lines.append(",".join(f'"{v}"' for v in row))

            csv_content = "\n".join(lines) + "\n"
            result[f"{name}.csv"] = csv_content.encode("utf-8")

        return result

    def create_from_template(
        self,
        template_path: str,
        sheets: list[dict],
        metadata: Optional[dict] = None,
    ) -> bytes:
        """Create a workbook using an existing .xlsx file as template.

        The template preserves formatting, formulas, charts, and styles.
        Data sheets from the request are added as new sheets (or replace
        existing sheets with matching names).

        Args:
            template_path: Path to the .xlsx template file.
            sheets: List of sheet data to populate.
            metadata: Optional document metadata.

        Returns:
            Workbook content as bytes.
        """
        validate_sheet_data(sheets)
        wb = load_workbook(template_path)

        # Apply metadata
        if metadata:
            self._apply_metadata(wb, metadata)

        # Build a set of sheet names to replace
        sheet_names = {s.get("name", "") for s in sheets}

        # Remove sheets that will be replaced
        for name in sheet_names:
            if name in wb.sheetnames:
                del wb[name]

        # Add/replace sheets with data
        for sheet_data in sheets:
            self._add_sheet(wb, sheet_data)

        return self._save_workbook(wb)
