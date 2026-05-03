"""Style application utilities for applying themes to documents."""

from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart

from src.styles.themes import Theme, get_theme
from src.utils.colors import ensure_argb_hex as _ensure_hex


class StyleApplier:
    """Applies theme styles to openpyxl workbooks."""

    def __init__(self, theme_name: str = "corporate"):
        self.theme = get_theme(theme_name)

    def apply_header_style(self, row, sheet):
        """Apply header styling to a row of cells."""
        colors = self.theme.colors
        fonts = self.theme.fonts

        header_font = Font(
            name=fonts.heading,
            size=12,
            bold=True,
            color=_ensure_hex(colors.header_text),
        )
        header_fill = PatternFill(
            start_color=_ensure_hex(colors.header_bg),
            end_color=_ensure_hex(colors.header_bg),
            fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(color=_ensure_hex(colors.border)),
            right=Side(color=_ensure_hex(colors.border)),
            bottom=Side(color=_ensure_hex(colors.border)),
        )

        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def apply_data_style(self, row, sheet, is_alternate: bool = False):
        """Apply data row styling."""
        colors = self.theme.colors
        fonts = self.theme.fonts

        data_font = Font(
            name=fonts.body,
            size=fonts.body_size,
            color=_ensure_hex(colors.text),
        )

        fill_color = colors.table_alt_row if is_alternate else colors.background
        data_fill = PatternFill(
            start_color=_ensure_hex(fill_color),
            end_color=_ensure_hex(fill_color),
            fill_type="solid",
        )
        data_alignment = Alignment(vertical="center")
        thin_border = Border(
            left=Side(color=_ensure_hex(colors.border)),
            right=Side(color=_ensure_hex(colors.border)),
            bottom=Side(color=_ensure_hex(colors.border)),
        )

        for cell in row:
            cell.font = data_font
            cell.fill = data_fill
            cell.alignment = data_alignment
            cell.border = thin_border

    def auto_fit_columns(self, sheet, width_multiplier: float = 1.2):
        """Auto-fit column widths based on content."""
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass
            adjusted_width = min((max_length + 2) * width_multiplier, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def create_chart(self, chart_type: str, title: str = "Chart"):
        """Create a chart based on type."""
        chart_types = {
            "bar": BarChart,
            "column": BarChart,
            "line": LineChart,
            "pie": PieChart,
            "area": AreaChart,
        }

        ChartClass = chart_types.get(chart_type, BarChart)
        chart = ChartClass()
        chart.title = title
        chart.width = 20
        chart.height = 12

        if chart_type == "column":
            chart.style = 2
            chart.type = "col"
        elif chart_type == "bar":
            chart.style = 10
            chart.type = "bar"

        return chart